# -*- coding: utf-8 -*-
"""Generate data.js (PROMPT_DATA) from RP.xlsx Sheet1 using the 3-level
module -> categories -> items structure (mirroring analysis-data.js)."""
import openpyxl
import json
import re

SRC = r'D:/weixinliaotianjilu/xwechat_files/wxid_5c7f98u5iv7521_64ca/msg/file/2026-07/RP.xlsx'
OUT = r'G:\SYF_Project\提示词网站\data.js'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Sheet1']

# Dimension metadata keyed by Excel dimension number (0-12)
DIM_META = {
    0:  {'title': '约束类型', 'role': None, 'multi': True,
         'desc': '控制 AI 对建筑几何、比例与材质的约束等级（轻度/中度/高度）',
         'cat_roles': {'轻度': 'constraint_light', '中度': 'constraint_medium', '高度': 'constraint_high'}},
    1:  {'title': '建筑类型', 'role': 'subject', 'multi': False,
         'desc': '指定渲染的建筑类型（医院 / 公建 / 文化 / 居住等）'},
    2:  {'title': '外立面/表皮材质', 'role': 'material', 'multi': False,
         'desc': '建筑外立面与表皮的材质系统'},
    3:  {'title': '材质材料', 'role': 'material', 'multi': False,
         'desc': '具体建筑材料与质感表现'},
    4:  {'title': '建筑风格/设计语言', 'role': 'style', 'multi': False,
         'desc': '建筑风格与设计语言'},
    5:  {'title': '色彩方案', 'role': 'color', 'multi': False,
         'desc': '整体色彩方案与心理效应'},
    6:  {'title': '光线/时间', 'role': 'light', 'multi': False,
         'desc': '光照条件与时段氛围'},
    7:  {'title': '天气氛围', 'role': 'weather', 'multi': False,
         'desc': '天气与大气环境'},
    8:  {'title': '视角/镜头', 'role': 'view', 'multi': False,
         'desc': '拍摄视角与镜头语言'},
    9:  {'title': '环境/周边', 'role': 'context', 'multi': False,
         'desc': '场地周边环境语境'},
    10: {'title': '氛围/情绪', 'role': 'mood', 'multi': False,
         'desc': '空间氛围与情绪基调'},
    11: {'title': '图像质量', 'role': 'quality', 'multi': False,
         'desc': '渲染画质与质量参数'},
    12: {'title': '自然元素/景观', 'role': 'landscape', 'multi': False,
         'desc': '自然元素与景观设计'},
}

# Ordered module ids 1..13 mapped from Excel dimension number
ORDER = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]


def js_str(s):
    return json.dumps(s, ensure_ascii=False)


def clean(v):
    if v is None:
        return ''
    return str(v).strip()


modules = []

current_dim = None
header_found = False
cur_cat = None          # for multi-cat dims
cur_cat_role = None
cur_cat_items = None
single_cat_items = None
single_cat_role = None
single_cat_title = None

rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))


def flush_dim():
    global cur_cat, cur_cat_items, single_cat_items
    if current_dim is None:
        return
    meta = DIM_META[current_dim]
    if meta['multi']:
        # categories already built in cat_map (ordered)
        cats = []
        for cname in ['轻度', '中度', '高度']:
            if cname in cat_map:
                cats.append({'title': cname, 'role': meta['cat_roles'][cname],
                             'items': cat_map[cname]})
        modules.append({'id': str(len(modules) + 1), 'title': meta['title'],
                        'desc': meta['desc'], 'categories': cats})
    else:
        if single_cat_items is not None:
            modules.append({'id': str(len(modules) + 1), 'title': meta['title'],
                            'desc': meta['desc'],
                            'categories': [{'title': single_cat_title,
                                            'role': single_cat_role,
                                            'items': single_cat_items}]})


cat_map = {}

for row in rows:
    a = clean(row[0]) if len(row) > 0 else ''
    b = clean(row[1]) if len(row) > 1 else ''
    c = clean(row[2]) if len(row) > 2 else ''
    d = clean(row[3]) if len(row) > 3 else ''
    e = clean(row[4]) if len(row) > 4 else ''

    # Detect new dimension header in col A
    m = re.match(r'^维度\s*(\d+)\s*[：:]', a)
    if m:
        flush_dim()
        current_dim = int(m.group(1))
        header_found = False
        cur_cat = None
        cur_cat_items = None
        single_cat_items = None
        cat_map = {}
        continue

    if current_dim is None:
        continue

    meta = DIM_META[current_dim]

    # Detect the keyword header row (col C contains 关键词)
    if '关键词' in c:
        header_found = True
        if not meta['multi']:
            single_cat_title = meta['title']
            single_cat_role = meta['role']
            single_cat_items = []
        continue

    # Data row: requires b or c content
    if not (b or c):
        continue

    # For dimensions without a keyword-header row, initialize the single
    # category on the first encountered data row.
    if not meta['multi'] and single_cat_items is None:
        single_cat_title = meta['title']
        single_cat_role = meta['role']
        single_cat_items = []

    if meta['multi']:
        # 维度0: col B = category grouping (轻度/中度/高度), item = colC(EN)+colD(CN)
        if b and b in meta['cat_roles']:
            cur_cat = b
            if cur_cat not in cat_map:
                cat_map[cur_cat] = []
        if (c or d):
            label = d if d else b
            item = {'label': label, 'en': c, 'cn': d if d else c}
            cat_map.setdefault(cur_cat or '轻度', []).append(item)
    else:
        # Non-multi: col B = CN label, col C = EN, col D = CN detail, col E = extra/effect
        if not (b or c):
            continue
        if current_dim == 8:  # 视角/镜头: col D=焦距, col E=效果
            item = {'label': b, 'en': c, 'cn': b}
            if d:
                item['extra'] = '焦距: ' + d
            if e:
                item['effect'] = e
        else:
            cn = d if d else b
            item = {'label': b, 'en': c, 'cn': cn}
            if e:
                item['extra'] = e
        single_cat_items.append(item)

flush_dim()

# Build JS
lines = []
lines.append('/**')
lines.append(' * 医院建筑 AI 效果图提示词数据')
lines.append(' * 三级层级结构：模块(1-13) → 类别 → 词条')
lines.append(' * 结构参考 analysis-data.js（分析图），与效果图模式共享模块→类别→词条模型')
lines.append(' */')
lines.append('const PROMPT_DATA = [')

for i, mod in enumerate(modules):
    comma = ',' if i < len(modules) - 1 else ''
    lines.append('  {')
    lines.append('    id: ' + js_str(mod['id']) + ',')
    lines.append('    title: ' + js_str(mod['title']) + ',')
    lines.append('    desc: ' + js_str(mod['desc']) + ',')
    lines.append('    categories: [')
    for j, cat in enumerate(mod['categories']):
        ccomma = ',' if j < len(mod['categories']) - 1 else ''
        lines.append('      {')
        lines.append('        title: ' + js_str(cat['title']) + ',')
        lines.append('        role: ' + js_str(cat['role']) + ',')
        lines.append('        items: [')
        for k, item in enumerate(cat['items']):
            icomma = ',' if k < len(cat['items']) - 1 else ''
            # build item literal
            kv = []
            kv.append('label: ' + js_str(item['label']))
            kv.append('en: ' + js_str(item['en']))
            kv.append('cn: ' + js_str(item['cn']))
            if 'extra' in item:
                kv.append('extra: ' + js_str(item['extra']))
            if 'effect' in item:
                kv.append('effect: ' + js_str(item['effect']))
            lines.append('          { ' + ', '.join(kv) + ' }' + icomma)
        lines.append('        ]')
        lines.append('      }' + ccomma)
    lines.append('    ]')
    lines.append('  }' + comma)

lines.append('];')
lines.append('')
lines.append("if (typeof module !== 'undefined' && module.exports) {")
lines.append('    module.exports = PROMPT_DATA;')
lines.append('}')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines) + '\n')

# Summary
print('Modules generated:', len(modules))
for mod in modules:
    total = sum(len(c['items']) for c in mod['categories'])
    print('  Module', mod['id'], mod['title'], '->', len(mod['categories']), 'categories,', total, 'items')
print('TOTAL items:', sum(sum(len(c['items']) for c in m['categories']) for m in modules))
